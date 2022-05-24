import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
products_under_10 = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # exercises1
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        product_per_supplier[supplier_name] = 1

    # exercises2
    if supplier_name in total_value_per_supplier:
        current_total_values = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_values + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # exercises3
    if inventory < 10 :
        products_under_10[int(product_num)] = int(inventory)

    # exercises4
    inventory_price.value = inventory * price

inv_file.save("inventory_finished.xlsx")

